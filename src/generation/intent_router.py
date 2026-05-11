"""Routage intelligent des requêtes : découverte de schéma + classification LLM.

Remplace les listes de mots-clés hard-codés du pipeline par :
1. SchemaDiscovery : scan automatique des fichiers du dossier docs (colonnes, échantillons)
2. IntentRouter   : appel LLM léger qui retourne {intent, source, column, exhaustive}
"""
import json
import logging
import re
from collections import OrderedDict
from pathlib import Path
from typing import Dict, List, Optional

from .llm import HFClient
from ..ingestion.loader import _detect_header_row

logger = logging.getLogger(__name__)


# ── SCHEMA DISCOVERY ─────────────────────────────────────────────────────────

class SchemaDiscovery:
    """Scanne le dossier de documents et extrait, pour chaque fichier :
       - les colonnes (Excel) ou un flag is_doc=True (Word/PDF/etc.)
       - quelques échantillons de valeurs par colonne (max 3, courts)
    Le résultat sert à injecter le schéma dans le prompt du classifieur."""

    EXCEL_EXTS = {".xlsx", ".xls"}
    DOC_EXTS = {".docx", ".doc", ".pdf", ".txt", ".md", ".html", ".htm"}

    def __init__(self, docs_dir: str, max_samples: int = 3, max_sample_len: int = 40):
        self.docs_dir = docs_dir
        self.max_samples = max_samples
        self.max_sample_len = max_sample_len

    def scan(self) -> Dict[str, dict]:
        schema: Dict[str, dict] = {}
        if not Path(self.docs_dir).exists():
            logger.warning("SchemaDiscovery: docs_dir introuvable: %s", self.docs_dir)
            return schema

        for file in sorted(Path(self.docs_dir).rglob("*")):
            if not file.is_file():
                continue
            ext = file.suffix.lower()
            stem = self._normalize_stem(file.name)
            if ext in self.EXCEL_EXTS:
                entry = self._scan_excel(file)
                if entry:
                    schema[stem] = entry
            elif ext in self.DOC_EXTS:
                schema[stem] = {"columns": [], "samples": {}, "is_doc": True,
                                "filename": file.name}

        logger.info("SchemaDiscovery: %d sources détectées (%s)",
                    len(schema), ", ".join(schema.keys()))
        return schema

    @staticmethod
    def _normalize_stem(fname: str) -> str:
        """SERVICE (1).xlsx → SERVICE  ;  KAM_Formations_GTP.xlsx → KAM_FORMATIONS_GTP"""
        stem = fname.rsplit(".", 1)[0] if "." in fname else fname
        stem = stem.upper().strip()
        stem = re.sub(r"\s*\(\d+\)\s*$", "", stem)
        stem = re.sub(r"\s*_\d+\s*$", "", stem)
        return stem.strip()

    def _scan_excel(self, path: Path) -> Optional[dict]:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
            ws = wb.active
            header_row = _detect_header_row(ws)
            headers: List[str] = []
            for cell in ws[header_row]:
                headers.append(str(cell.value).strip() if cell.value else "")
            headers = [h for h in headers if h]

            samples: Dict[str, List[str]] = {h: [] for h in headers}
            scanned = 0
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                scanned += 1
                if scanned > 200:  # limite pour gros Excel
                    break
                for h, v in zip(headers, row):
                    if v is None or len(samples[h]) >= self.max_samples:
                        continue
                    s = str(v).strip()[:self.max_sample_len]
                    if s and s not in samples[h]:
                        samples[h].append(s)
            wb.close()
            return {
                "columns": headers,
                "samples": {h: samples[h] for h in headers if samples[h]},
                "is_doc": False,
                "filename": path.name,
            }
        except Exception as e:
            logger.warning("SchemaDiscovery: échec scan %s: %s", path.name, e)
            return None


# ── INTENT ROUTER ─────────────────────────────────────────────────────────────

_SYSTEM_PROMPT = (
    "Tu es un classifieur de requêtes en français. "
    "Tu retournes UNIQUEMENT un objet JSON valide, rien d'autre."
)

_PROMPT_TEMPLATE = """Classifie la question utilisateur en JSON selon le schéma ci-dessous.

Sources disponibles (chaque ligne = 1 table, avec ses colonnes et des exemples de valeurs):
{schema_block}

Format JSON attendu (strict, une seule ligne):
{{"intent":"list|detail|qa","source":"<NOM_SOURCE_OU_null>","column":"<NOM_COLONNE_OU_null>","exhaustive":true|false,"filter":null}}

Règles:
- "intent":"list" si la question demande une énumération ("liste", "tous", "donne-moi les", "quels sont").
- "intent":"detail" si la question demande une explication ("explique", "détails").
- "intent":"qa" pour toute question ciblée ("qui est X", "combien").
- "source": EXACTEMENT un nom de table parmi la liste, ou null si vraiment aucune ne convient.
- "column": EXACTEMENT un nom de colonne de la table choisie. Regarde les exemples de valeurs pour identifier la bonne colonne. null si la question vise toute la ligne.
- "exhaustive":true si intent="list", sinon false.
- "filter": dict {{"NOM_COLONNE_REEL":"valeur"}} pour les contraintes. UTILISE LE VRAI NOM DE COLONNE de la table (vu dans le schéma), JAMAIS un nom inventé.

Exemples (basés sur le schéma ci-dessus):
Q: "Donne-moi la liste des services"
JSON: {{"intent":"list","source":"SERVICE","column":null,"exhaustive":true,"filter":null}}

Q: "Quels sont les directeurs ?"
JSON: {{"intent":"list","source":"DIRECTION","column":null,"exhaustive":true,"filter":null}}

Q: "Liste des chefs de département"
JSON: {{"intent":"list","source":"DEPARTEMENT","column":null,"exhaustive":true,"filter":null}}

Q: "Services de la direction DRH"
JSON: {{"intent":"list","source":"SERVICE","column":null,"exhaustive":true,"filter":{{"SHORT_LIBELLE_DIRECTION":"DRH"}}}}

Q: "Postes dans l'activité électricité"
JSON: {{"intent":"list","source":"POSTE","column":null,"exhaustive":true,"filter":{{"LIBELLE_ACTIVITE":"électricité"}}}}

Q: "Qui est le chef du département DRH ?"
JSON: {{"intent":"qa","source":"DEPARTEMENT","column":null,"exhaustive":false,"filter":{{"SHORT_LIBELLE_DIRECTION":"DRH"}}}}

Q: "Combien de directions ?"
JSON: {{"intent":"list","source":"DIRECTION","column":null,"exhaustive":true,"filter":null}}

Question: {question}
JSON:"""


class IntentRouter:
    """Classifie une question utilisateur via le LLM en s'appuyant sur le schéma découvert."""

    def __init__(self, llm: HFClient, schema: Dict[str, dict], cache_size: int = 256):
        self.llm = llm
        self.schema = schema
        self.cache_size = cache_size
        self._cache: "OrderedDict[str, dict]" = OrderedDict()
        self._schema_block = self._build_schema_block(schema)

    # ── public ────────────────────────────────────────────────────────────
    def classify(self, question: str) -> dict:
        key = self._normalize_question(question)
        if key in self._cache:
            self._cache.move_to_end(key)
            return self._cache[key]

        prompt = _PROMPT_TEMPLATE.format(
            schema_block=self._schema_block,
            question=question.strip(),
        )
        try:
            raw = self.llm.generate(
                prompt=prompt,
                system=_SYSTEM_PROMPT,
                temperature=0.0,
                max_tokens=120,
            )
        except Exception as e:
            logger.warning("IntentRouter: appel LLM échoué (%s) — fallback qa", e)
            return self._fallback()

        parsed = self._parse_json(raw)
        result = self._validate(parsed) if parsed else self._fallback()
        logger.info("IntentRouter: intent=%s source=%s column=%s exhaustive=%s conf=%s",
                    result["intent"], result["source"], result["column"],
                    result["exhaustive"], result["confidence"])

        self._cache[key] = result
        if len(self._cache) > self.cache_size:
            self._cache.popitem(last=False)
        return result

    # ── private ──────────────────────────────────────────────────────────
    @staticmethod
    def _normalize_question(q: str) -> str:
        q = q.lower().strip()
        for src, dst in [("é","e"),("è","e"),("ê","e"),("ë","e"),
                         ("à","a"),("â","a"),("ä","a"),
                         ("î","i"),("ï","i"),("ô","o"),("ö","o"),
                         ("ù","u"),("û","u"),("ü","u"),("ç","c")]:
            q = q.replace(src, dst)
        return re.sub(r"\s+", " ", q)

    def _build_schema_block(self, schema: Dict[str, dict]) -> str:
        """Expose chaque source avec colonnes + 2-3 valeurs d'exemple par colonne.

        Indispensable pour que le LLM devine la sémantique des colonnes
        techniques (ex: 'CHANTIER'='Service Informatique' → c'est un nom de
        service, pas un chantier de BTP). Sans samples, le routing échoue.
        """
        if not schema:
            return "(aucune source disponible)"
        lines = []
        for name, info in schema.items():
            if info.get("is_doc"):
                lines.append(f"* {name} (document texte) — descriptions, explications")
                continue
            cols = info.get("columns", [])
            samples = info.get("samples", {})
            row_count = info.get("row_count", "?")

            # En-tête : nom + nb lignes
            lines.append(f"* {name} ({row_count} lignes) :")

            # Une ligne par colonne, avec 2-3 exemples
            for col in cols:
                vals = samples.get(col, [])
                if vals:
                    sample_str = ", ".join(f'"{v}"' for v in vals[:3])
                    lines.append(f"    - {col} → ex: {sample_str}")
                else:
                    lines.append(f"    - {col}")
        return "\n".join(lines)

    def _parse_json(self, raw: str) -> Optional[dict]:
        if not raw:
            return None
        # 1. tenter d'isoler le 1er { ... } de la réponse
        match = re.search(r"\{.*\}", raw, flags=re.DOTALL)
        candidate = match.group(0) if match else raw.strip()
        for attempt in (candidate, self._repair_json(candidate)):
            try:
                return json.loads(attempt)
            except (json.JSONDecodeError, TypeError):
                continue
        return None

    @staticmethod
    def _repair_json(s: str) -> str:
        # single → double quotes (basique), retire trailing commas
        s = re.sub(r"'", '"', s)
        s = re.sub(r",\s*([}\]])", r"\1", s)
        return s

    def _validate(self, data: dict) -> dict:
        intent = str(data.get("intent", "qa")).lower()
        if intent not in ("list", "detail", "qa"):
            intent = "qa"

        source = data.get("source")
        if source in ("null", "None", ""):
            source = None
        if source and source not in self.schema:
            # tentative de match insensible à la casse / au stem
            up = str(source).upper()
            source = up if up in self.schema else None

        column = data.get("column")
        if column in ("null", "None", ""):
            column = None
        if column and source and not self.schema[source].get("is_doc"):
            cols_upper = {c.upper(): c for c in self.schema[source].get("columns", [])}
            column = cols_upper.get(str(column).upper())

        exhaustive = bool(data.get("exhaustive", False)) or intent == "list"

        filt = data.get("filter")
        if not isinstance(filt, dict) or not filt:
            filt = None
        elif source:  # normaliser les clés du filtre sur les colonnes connues
            cols_upper = {c.upper(): c for c in self.schema[source].get("columns", [])}
            filt = {cols_upper.get(str(k).upper(), k): v for k, v in filt.items()}

        return {
            "intent": intent,
            "source": source,
            "column": column,
            "exhaustive": exhaustive,
            "filter": filt,
            "confidence": "high",
        }

    @staticmethod
    def _fallback() -> dict:
        return {
            "intent": "qa", "source": None, "column": None,
            "exhaustive": False, "filter": None, "confidence": "low",
        }
