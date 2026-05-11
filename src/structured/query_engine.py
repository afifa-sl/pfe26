"""StructuredQueryEngine — interroge les fichiers Excel via DuckDB en mémoire.

Architecture :
    1. Au démarrage, scanne docs_dir et charge chaque .xlsx dans une table DuckDB.
    2. Auto-détection de la ligne d'en-tête (réutilise loader._detect_header_row)
       pour gérer les Excel avec titre/métadonnées avant le tableau.
    3. Expose des méthodes haut niveau (list_values, filter_rows, count_rows)
       qui construisent du SQL paramétré — JAMAIS de concaténation directe de
       valeurs utilisateur dans les requêtes (protection injection SQL).

Pourquoi DuckDB plutôt que pandas / SQLite :
    - Lecture directe des .xlsx via l'extension `excel` (pas de conversion).
    - SQL standard, performances excellentes sur dizaines de milliers de lignes.
    - In-memory par défaut, zéro fichier à gérer.
"""
from __future__ import annotations

import logging
import re
import unicodedata
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


class StructuredQueryEngine:
    EXCEL_EXTS = {".xlsx", ".xls"}

    def __init__(self, docs_dir: str):
        import duckdb
        self._duckdb = duckdb
        self.conn = duckdb.connect(":memory:")
        self.tables: Dict[str, Dict[str, Any]] = {}
        self.last_warnings: List[str] = []
        self.docs_dir = docs_dir
        self._load_all(docs_dir)

    # ── Chargement ────────────────────────────────────────────────────────

    def _load_all(self, docs_dir: str) -> None:
        if not Path(docs_dir).exists():
            logger.warning("StructuredQueryEngine: docs_dir introuvable: %s", docs_dir)
            return
        loaded = 0
        for file in sorted(Path(docs_dir).rglob("*")):
            if not file.is_file() or file.suffix.lower() not in self.EXCEL_EXTS:
                continue
            try:
                self._load_excel(file)
                loaded += 1
            except Exception as e:
                logger.warning("  ⚠ Skip %s: %s", file.name, e)
        logger.info("StructuredQueryEngine: %d table(s) chargée(s) en DuckDB (%s)",
                    loaded, ", ".join(self.tables.keys()) or "—")

    def _load_excel(self, path: Path) -> None:
        from ..ingestion.loader import _detect_header_row
        import openpyxl

        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        try:
            ws = wb.active
            header_row = _detect_header_row(ws)
            headers: List[str] = []
            for cell in ws[header_row]:
                headers.append(str(cell.value).strip() if cell.value else "")

            keep_idx = [i for i, h in enumerate(headers) if h]
            sql_headers = [self._sql_ident(headers[i]) for i in keep_idx]
            if not sql_headers:
                logger.warning("  ⚠ %s: aucune colonne valide", path.name)
                return

            seen: Dict[str, int] = {}
            unique_headers: List[str] = []
            for h in sql_headers:
                if h in seen:
                    seen[h] += 1
                    unique_headers.append(f"{h}_{seen[h]}")
                else:
                    seen[h] = 0
                    unique_headers.append(h)

            rows: List[List[Optional[str]]] = []
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                values = [row[i] if i < len(row) else None for i in keep_idx]
                values = [None if v is None else str(v).strip() for v in values]
                if any(v for v in values):
                    rows.append(values)
        finally:
            wb.close()

        if not rows:
            logger.warning("  ⚠ %s: aucune donnée", path.name)
            return

        table_name = self._normalize_stem(path.name)
        sql_table = self._sql_ident(table_name)

        cols_def = ", ".join(f'"{c}" VARCHAR' for c in unique_headers)
        self.conn.execute(f'CREATE OR REPLACE TABLE "{sql_table}" ({cols_def})')

        placeholders = ", ".join(["?"] * len(unique_headers))
        self.conn.executemany(
            f'INSERT INTO "{sql_table}" VALUES ({placeholders})',
            rows,
        )

        technical_cols = self._detect_technical_columns(sql_table, unique_headers, len(rows))

        self.tables[table_name] = {
            "filename": path.name,
            "columns": unique_headers,
            "technical_columns": technical_cols,
            "user_columns": [c for c in unique_headers if c not in technical_cols],
            "row_count": len(rows),
            "sql_table": sql_table,
        }
        if technical_cols:
            logger.info("  ✓ %s → table %s (%d lignes, %d colonnes ; techniques: %s)",
                        path.name, sql_table, len(rows), len(unique_headers),
                        ", ".join(sorted(technical_cols)))
        else:
            logger.info("  ✓ %s → table %s (%d lignes, %d colonnes)",
                        path.name, sql_table, len(rows), len(unique_headers))

    def _detect_technical_columns(self, sql_table: str, columns: List[str], row_count: int) -> set:
        if row_count == 0:
            return set()
        sample_size = min(row_count, 200)
        techincal_name_hints = {"ID", "CODE", "CD"}
        technical: set = set()
        for col in columns:
            up = col.upper()
            if up in techincal_name_hints or up.startswith(("CD_", "ID_", "NUM_")):
                technical.add(col)
                continue
            try:
                rows = self.conn.execute(
                    f'SELECT "{col}" FROM "{sql_table}" '
                    f'WHERE "{col}" IS NOT NULL AND TRIM("{col}") <> \'\' '
                    f'LIMIT {sample_size}'
                ).fetchall()
            except Exception:
                continue
            if not rows:
                continue
            vals = [str(r[0]).strip() for r in rows]
            n = len(vals)
            numeric = sum(1 for v in vals if v.replace(".", "").replace(",", "").replace("-", "").isdigit())
            short_alphanum = sum(1 for v in vals if len(v) <= 6 and v.replace("_", "").replace("-", "").isalnum())
            if numeric / n > 0.7:
                technical.add(col)
            elif short_alphanum / n > 0.7 and numeric / n > 0.3:
                technical.add(col)
        return technical

    # ── Helpers normalisation ─────────────────────────────────────────────

    @staticmethod
    def _normalize_stem(fname: str) -> str:
        stem = fname.rsplit(".", 1)[0] if "." in fname else fname
        stem = stem.upper().strip()
        stem = re.sub(r"\s*\(\d+\)\s*$", "", stem)
        stem = re.sub(r"\s*_\d+\s*$", "", stem)
        return stem.strip()

    @staticmethod
    def _sql_ident(name: str) -> str:
        nfd = unicodedata.normalize("NFD", name)
        ascii_name = "".join(c for c in nfd if unicodedata.category(c) != "Mn")
        cleaned = re.sub(r"[^A-Za-z0-9_]+", "_", ascii_name).strip("_").upper()
        if not cleaned:
            cleaned = "COL"
        if cleaned[0].isdigit():
            cleaned = "C_" + cleaned
        return cleaned

    @staticmethod
    def _fold(text: str) -> str:
        nfd = unicodedata.normalize("NFD", str(text))
        ascii_text = "".join(c for c in nfd if unicodedata.category(c) != "Mn")
        return ascii_text.lower().strip()

    def _resolve_column(self, table: str, column: Optional[str]) -> Optional[str]:
        if not column:
            return None
        cols = self.tables[table]["columns"]
        target = self._fold(column)
        target_sql = self._sql_ident(column)
        for c in cols:
            if self._fold(c) == target or c == target_sql:
                return c
        return None

    # ── API publique ──────────────────────────────────────────────────────

    def has_table(self, name: str) -> bool:
        return self._normalize_stem(name) in self.tables \
            or name in self.tables

    def schema(self) -> Dict[str, Dict[str, Any]]:
        return {
            name: {
                "columns": info["columns"],
                "filename": info["filename"],
                "row_count": info["row_count"],
                "is_doc": False,
            }
            for name, info in self.tables.items()
        }

    def samples(self, table: str, max_per_col: int = 3, max_len: int = 40) -> Dict[str, List[str]]:
        table = self._normalize_stem(table) if table not in self.tables else table
        if table not in self.tables:
            return {}
        sql_table = self.tables[table]["sql_table"]
        out: Dict[str, List[str]] = {}
        for col in self.tables[table]["columns"]:
            try:
                rows = self.conn.execute(
                    f'SELECT DISTINCT "{col}" FROM "{sql_table}" '
                    f'WHERE "{col}" IS NOT NULL LIMIT {max_per_col}'
                ).fetchall()
                vals = [str(r[0])[:max_len] for r in rows if r[0]]
                if vals:
                    out[col] = vals
            except Exception:
                continue
        return out

    def list_values(
        self,
        table: str,
        column: Optional[str] = None,
        filters: Optional[Dict[str, str]] = None,
        distinct: bool = True,
        hide_technical: bool = True,
    ) -> List[Dict[str, Any]]:
        self.last_warnings = []
        table = self._normalize_stem(table) if table not in self.tables else table
        if table not in self.tables:
            return []
        sql_table = self.tables[table]["sql_table"]
        meta_filename = self.tables[table]["filename"]
        all_cols = self.tables[table]["columns"]

        resolved_filters: List[tuple] = []
        dropped_filters: List[str] = []
        for fkey, fval in (filters or {}).items():
            sql_col_filter = self._resolve_column(table, fkey)
            if sql_col_filter and fval:
                resolved_filters.append((sql_col_filter, str(fval)))
            else:
                dropped_filters.append(f"{fkey}={fval!r}")
        if dropped_filters:
            msg = (f"Filtre(s) ignoré(s) : {', '.join(dropped_filters)} "
                   f"— colonne(s) absente(s) de la table {table} "
                   f"(colonnes disponibles : {', '.join(all_cols)}).")
            logger.warning("  ⚠ %s", msg)
            self.last_warnings.append(msg)

        sql_col = self._resolve_column(table, column) if column else None

        if sql_col:
            select_clause = f'"{sql_col}"'
        else:
            if hide_technical:
                cols = self.tables[table].get("user_columns") or self.tables[table]["columns"]
            else:
                cols = self.tables[table]["columns"]
            select_clause = ", ".join(f'"{c}"' for c in cols)

        distinct_kw = "DISTINCT " if distinct else ""

        where_parts = []
        params: List[str] = []
        if sql_col:
            where_parts.append(f'"{sql_col}" IS NOT NULL')
            where_parts.append(f'TRIM("{sql_col}") <> \'\'')
        for col, val in resolved_filters:
            where_parts.append(
                f'strip_accents(LOWER(CAST("{col}" AS VARCHAR))) '
                f'LIKE strip_accents(LOWER(?))'
            )
            params.append(f"%{val}%")

        where_clause = (" WHERE " + " AND ".join(where_parts)) if where_parts else ""
        sql = f'SELECT {distinct_kw}{select_clause} FROM "{sql_table}"{where_clause}'

        try:
            cursor = self.conn.execute(sql, params)
            rows = cursor.fetchall()
            col_names = [d[0] for d in cursor.description]
        except Exception as e:
            logger.warning("StructuredQueryEngine: requête SQL échouée (%s) — fallback lower", e)
            return self._list_values_fallback(table, column, filters, distinct)

        if not rows and resolved_filters:
            tokenized_where = []
            tokenized_params: List[str] = []
            if sql_col:
                tokenized_where.append(f'"{sql_col}" IS NOT NULL')
                tokenized_where.append(f'TRIM("{sql_col}") <> \'\'')
            had_tokens = False
            for col, val in resolved_filters:
                tokens = [t for t in re.split(r"\s+", val.strip()) if len(t) >= 3]
                tokens = [t[:5] if len(t) > 5 else t for t in tokens]
                for t in tokens:
                    tokenized_where.append(
                        f'strip_accents(LOWER(CAST("{col}" AS VARCHAR))) '
                        f'LIKE strip_accents(LOWER(?))'
                    )
                    tokenized_params.append(f"%{t}%")
                    had_tokens = True
            if had_tokens:
                fallback_sql = (
                    f'SELECT {distinct_kw}{select_clause} FROM "{sql_table}" '
                    f'WHERE ' + " AND ".join(tokenized_where)
                )
                try:
                    cursor = self.conn.execute(fallback_sql, tokenized_params)
                    rows = cursor.fetchall()
                    col_names = [d[0] for d in cursor.description]
                    if rows:
                        msg = (f"Aucun résultat pour le filtre exact, "
                               f"{len(rows)} résultat(s) en match approximatif "
                               f"(tokens: {tokenized_params}).")
                        logger.info("  ⟹ %s", msg)
                        self.last_warnings.append(msg)
                except Exception as e:
                    logger.warning("Fallback tokenisé échoué: %s", e)

        results: List[Dict[str, Any]] = []
        for row in rows:
            if sql_col:
                val = row[0]
                if val is None or not str(val).strip():
                    continue
                results.append({
                    "content": str(val).strip(),
                    "metadata": {"filename": meta_filename, "table": table, "column": sql_col},
                })
            else:
                pairs = [f"{c}: {v}" for c, v in zip(col_names, row) if v is not None and str(v).strip()]
                if not pairs:
                    continue
                content = f"[{table}] " + " | ".join(pairs)
                results.append({
                    "content": content,
                    "metadata": {"filename": meta_filename, "table": table},
                })

        logger.info("  ⟹ DuckDB: %d résultat(s) [table=%s, col=%s, filters=%s]",
                    len(results), table, sql_col or "ALL", filters or "-")
        return results

    def _list_values_fallback(self, table, column, filters, distinct, hide_technical=True):
        sql_table = self.tables[table]["sql_table"]
        meta_filename = self.tables[table]["filename"]
        cols = self.tables[table]["columns"]
        user_cols = (self.tables[table].get("user_columns") or cols) if hide_technical else cols
        rows = self.conn.execute(f'SELECT * FROM "{sql_table}"').fetchall()
        sql_col = self._resolve_column(table, column) if column else None

        resolved_filters = []
        for fkey, fval in (filters or {}).items():
            rc = self._resolve_column(table, fkey)
            if rc and fval:
                resolved_filters.append((rc, self._fold(fval)))

        results = []
        seen_distinct: set = set()
        for row in rows:
            kv = dict(zip(cols, row))
            ok = True
            for fc, fv in resolved_filters:
                actual = kv.get(fc)
                if actual is None or fv not in self._fold(str(actual)):
                    ok = False
                    break
            if not ok:
                continue
            if sql_col:
                val = kv.get(sql_col)
                if val is None or not str(val).strip():
                    continue
                key = self._fold(val)
                if distinct and key in seen_distinct:
                    continue
                seen_distinct.add(key)
                results.append({"content": str(val).strip(),
                                "metadata": {"filename": meta_filename, "table": table, "column": sql_col}})
            else:
                pairs = [f"{c}: {v}" for c, v in kv.items()
                         if c in user_cols and v is not None and str(v).strip()]
                if pairs:
                    results.append({"content": f"[{table}] " + " | ".join(pairs),
                                    "metadata": {"filename": meta_filename, "table": table}})
        return results

    def count_rows(self, table: str, filters: Optional[Dict[str, str]] = None) -> int:
        rows = self.list_values(table, column=None, filters=filters, distinct=False)
        return len(rows)

    # ── RECHARGEMENT DYNAMIQUE ────────────────────────────────────────────────

    def reload(self, docs_dir: Optional[str] = None) -> int:
        """Recharge tous les Excel depuis docs_dir dans DuckDB sans redémarrer l'app.

        À appeler après chaque ingestion (pipeline.ingest / ingest_documents).
        """
        target_dir = docs_dir or self.docs_dir
        self.conn.close()
        import duckdb
        self.conn = duckdb.connect(":memory:")
        self.tables.clear()
        self.last_warnings = []
        self._load_all(target_dir)
        logger.info("StructuredQueryEngine.reload(): %d table(s) — %s",
                    len(self.tables), ", ".join(self.tables.keys()) or "—")
        return len(self.tables)

    def get_primary_column(self, table: str) -> Optional[str]:
        """Détecte la colonne qui contient le NOM PRINCIPAL de chaque ligne.

        Stratégie en 3 passes (sans aucun nom de colonne hard-codé) :

        Passe 1 — Mots-clés sémantiques dans le NOM de la colonne
            On cherche des mots qui signifient "nom" ou "libellé" dans toutes
            les langues et conventions courantes des Excel RH/entreprise.
            Une colonne nommée CHANTIER, LIBELLE_SERVICE, NOM_DIRECTION, etc.
            obtient un bonus fort s'il y a des mots-clés dans son nom.
            → Évite de choisir FONCTION ou OBSERVATION à la place du libellé.

        Passe 2 — Analyse statistique des valeurs réelles
            Pour chaque colonne candidate, calcule :
              - ratio_unique : nb_distinct / nb_lignes  (proche de 1 = chaque
                ligne a son propre nom → c'est bien un identifiant descriptif)
              - avg_len : longueur moyenne des valeurs
            Score = ratio_unique × avg_len × bonus_semantique

        Passe 3 — Fallback
            Si aucun score positif, retourne la première colonne user.
        """
        import math

        table = self._normalize_stem(table) if table not in self.tables else table
        if table not in self.tables:
            return None

        sql_table  = self.tables[table]["sql_table"]
        user_cols  = self.tables[table].get("user_columns") or self.tables[table]["columns"]
        row_count  = self.tables[table]["row_count"]

        if not user_cols:
            return None
        if len(user_cols) == 1:
            return user_cols[0]

        # ── Passe 1 : bonus sémantique sur le NOM de la colonne ─────────────
        # Mots qui indiquent "libellé principal" dans les Excel algériens / FR
        # On teste des sous-chaînes pour couvrir LIBELLE_SERVICE, CHANTIER, etc.
        # On pénalise les colonnes qui décrivent un rôle/fonction, pas un nom.
        NAME_KEYWORDS   = [
            "LIBELLE", "NOM", "INTITULE", "DESIGNATION",
            "CHANTIER",   # convention Sonatrach pour nom de service
            "TITRE", "LABEL", "DESCRIPTION",
        ]
        PENALTY_KEYWORDS = [
            "FONCTION", "OBSERVATION", "STATUT", "STATUS",
            "PRENOM", "EMAIL", "TEL", "DATE", "MATRICULE",
            "GRADE", "CATEGORIE", "NIVEAU",
        ]

        def _semantic_bonus(col_name: str) -> float:
            up = col_name.upper()
            for kw in PENALTY_KEYWORDS:
                if kw in up:
                    return 0.0          # éliminé
            for kw in NAME_KEYWORDS:
                if kw in up:
                    return 2.0          # bonus fort
            # Le nom de la colonne == nom de la table → c'est souvent le libellé
            table_stem = table.upper().replace("_", "")
            col_stem   = up.replace("_", "")
            if table_stem in col_stem or col_stem in table_stem:
                return 1.5
            return 1.0                  # neutre

        # ── Passe 2 : analyse statistique ───────────────────────────────────
        best_col:   Optional[str] = None
        best_score: float = -1.0

        for col in user_cols:
            bonus = _semantic_bonus(col)
            if bonus == 0.0:
                continue               # colonne pénalisée → ignorée

            try:
                row = self.conn.execute(
                    f'SELECT COUNT(DISTINCT "{col}"), COUNT("{col}") '
                    f'FROM "{sql_table}" '
                    f'WHERE "{col}" IS NOT NULL AND TRIM("{col}") <> \'\''
                ).fetchone()
                distinct_count, non_null = (row[0], row[1]) if row else (0, 0)

                if distinct_count == 0:
                    continue

                sample = self.conn.execute(
                    f'SELECT "{col}" FROM "{sql_table}" '
                    f'WHERE "{col}" IS NOT NULL AND TRIM("{col}") <> \'\' '
                    f'LIMIT 30'
                ).fetchall()
                vals = [str(r[0]).strip() for r in sample if r[0]]

                if not vals:
                    continue

                avg_len = sum(len(v) for v in vals) / len(vals)

                # Colonnes avec valeurs très courtes → codes, pas des noms
                if avg_len < 4:
                    continue

                # ratio_unique proche de 1 → chaque ligne a un nom distinct
                ratio_unique = distinct_count / max(non_null, 1)

                # Score final
                score = bonus * ratio_unique * avg_len * math.log1p(distinct_count)

                logger.debug(
                    "  get_primary_column[%s] col=%s bonus=%.1f ratio=%.2f "
                    "avg_len=%.1f distinct=%d score=%.1f",
                    table, col, bonus, ratio_unique, avg_len, distinct_count, score,
                )

                if score > best_score:
                    best_score = score
                    best_col   = col

            except Exception as e:
                logger.debug("  get_primary_column: erreur col %s: %s", col, e)
                continue

        # ── Passe 3 : fallback ───────────────────────────────────────────────
        if best_col is None and user_cols:
            best_col = user_cols[0]

        logger.info("get_primary_column(%s) → %s", table, best_col)
        return best_col

    def close(self) -> None:
        try:
            self.conn.close()
        except Exception:
            pass
